import json
import os
import re
import sys
import zipfile
import xml.etree.ElementTree as ET
from collections import Counter, defaultdict
from datetime import datetime
from statistics import median

import openpyxl


path = sys.argv[1]
formula_book = openpyxl.load_workbook(path, read_only=True, data_only=False)
value_book = openpyxl.load_workbook(path, read_only=True, data_only=True)

result = {
    "path": os.path.abspath(path),
    "size_bytes": os.path.getsize(path),
    "sheets": [],
}

with zipfile.ZipFile(path) as archive:
    sheet_xml = archive.read("xl/worksheets/sheet1.xml")
    strings_xml = archive.read("xl/sharedStrings.xml")
    root = ET.fromstring(sheet_xml)
    strings_root = ET.fromstring(strings_xml)
    namespace = {"x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    shared_strings = [
        "".join(item.itertext())
        for item in strings_root.findall("x:si", namespace)
    ]
    dimension = root.find("x:dimension", namespace)
    raw_rows = root.findall(".//x:sheetData/x:row", namespace)
    raw_cells = root.findall(".//x:sheetData/x:row/x:c", namespace)
    decoded_rows = []
    for raw_row in raw_rows:
        decoded = {}
        for cell in raw_row.findall("x:c", namespace):
            coordinate = cell.attrib.get("r", "")
            column = re.match(r"[A-Z]+", coordinate).group(0)
            raw_value = cell.findtext("x:v", default=None, namespaces=namespace)
            if cell.attrib.get("t") == "s" and raw_value is not None:
                value = shared_strings[int(raw_value)]
            elif raw_value is None:
                value = None
            else:
                try:
                    value = float(raw_value)
                except ValueError:
                    value = raw_value
            decoded[column] = value
        decoded_rows.append({"row": int(raw_row.attrib["r"]), "cells": decoded})

    headers = decoded_rows[0]["cells"]
    records = []
    for item in decoded_rows[1:]:
        cells = item["cells"]
        record = {headers.get(column, column): value for column, value in cells.items()}
        record["_row"] = item["row"]
        records.append(record)

    type_counts = Counter()
    type_amounts = defaultdict(float)
    currencies = Counter()
    account1_counts = Counter()
    account2_counts = Counter()
    account_balances = defaultdict(float)
    account_inflows = defaultdict(float)
    account_outflows = defaultdict(float)
    annual = defaultdict(lambda: defaultdict(float))
    monthly = defaultdict(lambda: defaultdict(float))
    expense_categories = defaultdict(float)
    recent_expense_categories = defaultdict(float)
    recent_expense_secondary = defaultdict(float)
    unknown_types = Counter()
    parsed_dates = []

    for record in records:
        tx_type = str(record.get("类型") or "")
        amount = float(record.get("金额") or 0)
        currency = str(record.get("币种") or "")
        account1 = str(record.get("账户1") or "")
        account2 = str(record.get("账户2") or "")
        category = str(record.get("分类") or "")
        secondary = str(record.get("二级分类") or "")
        dt = datetime.fromisoformat(str(record.get("时间")))
        parsed_dates.append(dt)
        year = str(dt.year)
        month = dt.strftime("%Y-%m")

        type_counts[tx_type] += 1
        type_amounts[tx_type] += amount
        currencies[currency] += 1
        if account1:
            account1_counts[account1] += 1
        if account2:
            account2_counts[account2] += 1

        if tx_type in {"收入", "退款"}:
            account_balances[account1] += amount
            account_inflows[account1] += amount
        elif tx_type == "支出":
            account_balances[account1] -= amount
            account_outflows[account1] += amount
        elif tx_type in {"转账", "还款"}:
            account_balances[account1] -= amount
            account_outflows[account1] += amount
            account_balances[account2] += amount
            account_inflows[account2] += amount
        else:
            unknown_types[tx_type] += 1

        if tx_type == "收入":
            annual[year]["income"] += amount
            monthly[month]["income"] += amount
        elif tx_type == "支出":
            annual[year]["expense"] += amount
            monthly[month]["expense"] += amount
            expense_categories[category] += amount
        elif tx_type == "退款":
            annual[year]["refund"] += amount
            monthly[month]["refund"] += amount

    latest_dt = max(parsed_dates)
    latest_month_index = latest_dt.year * 12 + latest_dt.month - 1
    recent_start_index = latest_month_index - 11
    recent_months = []
    for idx in range(recent_start_index, latest_month_index + 1):
        recent_months.append(f"{idx // 12:04d}-{idx % 12 + 1:02d}")

    for record in records:
        tx_type = str(record.get("类型") or "")
        if tx_type != "支出":
            continue
        dt = datetime.fromisoformat(str(record.get("时间")))
        month = dt.strftime("%Y-%m")
        if month not in recent_months:
            continue
        amount = float(record.get("金额") or 0)
        recent_expense_categories[str(record.get("分类") or "")] += amount
        recent_expense_secondary[str(record.get("二级分类") or "")] += amount

    monthly_rows = []
    for month in recent_months:
        row = dict(monthly[month])
        row["month"] = month
        row["net_income_less_expense_plus_refund"] = (
            row.get("income", 0) - row.get("expense", 0) + row.get("refund", 0)
        )
        monthly_rows.append(row)

    expense_values = [row.get("expense", 0) for row in monthly_rows]
    income_values = [row.get("income", 0) for row in monthly_rows]
    summary = {
        "source_range": "账单!A1:S3708",
        "record_count": len(records),
        "earliest": min(parsed_dates).isoformat(sep=" "),
        "latest": latest_dt.isoformat(sep=" "),
        "headers": headers,
        "type_counts": dict(type_counts),
        "type_amounts": dict(sorted(type_amounts.items())),
        "currencies": dict(currencies),
        "unknown_types": dict(unknown_types),
        "derived_account_balances": dict(
            sorted(account_balances.items(), key=lambda item: item[1], reverse=True)
        ),
        "account_inflows": dict(sorted(account_inflows.items(), key=lambda item: item[1], reverse=True)),
        "account_outflows": dict(sorted(account_outflows.items(), key=lambda item: item[1], reverse=True)),
        "annual": {year: dict(values) for year, values in sorted(annual.items())},
        "trailing_12_months": monthly_rows,
        "trailing_12_median_monthly_expense": median(expense_values),
        "trailing_12_average_monthly_expense": sum(expense_values) / len(expense_values),
        "trailing_12_median_monthly_income": median(income_values),
        "all_time_expense_categories": dict(
            sorted(expense_categories.items(), key=lambda item: item[1], reverse=True)
        ),
        "trailing_12_expense_categories": dict(
            sorted(recent_expense_categories.items(), key=lambda item: item[1], reverse=True)
        ),
        "trailing_12_expense_secondary": dict(
            sorted(recent_expense_secondary.items(), key=lambda item: item[1], reverse=True)
        ),
        "account1_counts": dict(account1_counts),
        "account2_counts": dict(account2_counts),
    }

    if "--summary" in sys.argv:
        print(json.dumps(summary, ensure_ascii=False, default=str))
        sys.exit(0)
    result["raw_sheet1"] = {
        "dimension": dimension.attrib if dimension is not None else None,
        "row_count": len(raw_rows),
        "cell_count": len(raw_cells),
        "first_rows": [row.attrib for row in raw_rows[:20]],
        "first_cells": [
            {
                "attrib": cell.attrib,
                "value": cell.findtext("x:v", default=None, namespaces=namespace),
            }
            for cell in raw_cells[:80]
        ],
        "shared_string_count": len(shared_strings),
        "decoded_first_rows": decoded_rows[:100],
        "decoded_last_rows": decoded_rows[-20:],
        "xml_prefix": sheet_xml[:4000].decode("utf-8", "replace"),
    }

for formula_sheet in formula_book.worksheets:
    value_sheet = value_book[formula_sheet.title]
    nonempty_rows = []
    formula_cells = []
    for row_idx in range(1, (formula_sheet.max_row or 0) + 1):
        values = []
        has_value = False
        for col_idx in range(1, (formula_sheet.max_column or 0) + 1):
            formula_cell = formula_sheet.cell(row_idx, col_idx)
            value = formula_cell.value
            values.append(value)
            if value not in (None, ""):
                has_value = True
            if isinstance(value, str) and value.startswith("="):
                formula_cells.append(
                    {
                        "coordinate": formula_cell.coordinate,
                        "formula": value,
                        "cached_value": value_sheet.cell(row_idx, col_idx).value,
                    }
                )
        if has_value:
            nonempty_rows.append({"row": row_idx, "values": values})

    result["sheets"].append(
        {
            "name": formula_sheet.title,
            "max_row": formula_sheet.max_row,
            "max_column": formula_sheet.max_column,
            "merged_ranges": [
                str(item)
                for item in getattr(getattr(formula_sheet, "merged_cells", None), "ranges", [])
            ],
            "nonempty_rows": nonempty_rows,
            "formula_cells": formula_cells,
        }
    )

print(json.dumps(result, ensure_ascii=False, default=str))
